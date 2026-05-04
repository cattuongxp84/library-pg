# -*- coding: utf-8 -*-
import sys
import io
# Đảm bảo output UTF-8 trên Windows
if sys.platform == 'win32' and sys.stdout is not None:
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')
    except:
        pass

from flask import Flask, request, jsonify, render_template, redirect, url_for, flash, send_file, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
import requests
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from sqlalchemy import inspect, text, or_
from threading import Thread
import importlib

# OpenPyXL imports
if importlib.util.find_spec("openpyxl") is not None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
else:
    OPENPYXL_AVAILABLE = False

# Firebase imports
if importlib.util.find_spec("firebase_admin") is not None:
    import firebase_admin
    from firebase_admin import credentials, db as firebase_db
    FIREBASE_AVAILABLE = True
else:
    FIREBASE_AVAILABLE = False

import os, time, datetime, socket, logging, json, glob, webbrowser, pandas as pd

# Cấu hình file upload
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ==== Đường dẫn và khởi tạo thư mục ====
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

def ensure_directories():
    directories = [
        resource_path('static/uploads/ads'),
        resource_path('static/uploads/backgrounds'),
        resource_path('static/css'),
        resource_path('static/js'),
        resource_path('templates')
    ]
    for directory in directories:
        if not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)

ensure_directories()

# ==== Flask App & Config ====
app = Flask(__name__, 
            template_folder=resource_path('templates'),
            static_folder=resource_path('static'))
app.secret_key = 'your-very-secret-key'
CORS(app)
app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql+psycopg2://DPTIUH:libiuh2025@localhost:5432/lab_manager'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = resource_path('static/uploads/ads')
app.config['BACKGROUND_FOLDER'] = resource_path('static/uploads/backgrounds')
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['BACKGROUND_FOLDER'], exist_ok=True)
app.config['HEARTBEAT_INTERVAL'] = 5
app.config['HEARTBEAT_TIMEOUT'] = 15
app.config['EXTENDED_TIMEOUT'] = 45

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

client_apis = {}

# ==== Model ====
db = SQLAlchemy(app)

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.String(20), unique=True, nullable=False)
    name = db.Column(db.String(100))
    class_name = db.Column(db.String(50))
    khoa_vien = db.Column(db.String(100))
    password_hash = db.Column(db.String(255), nullable=False)
    plain_password = db.Column(db.String(128))
    created_at = db.Column(db.DateTime, default=datetime.datetime.now)
    sessions = db.relationship('Session', backref='user', lazy=True)
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class Computer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50))
    mac_address = db.Column(db.String(20))
    ip_address = db.Column(db.String(15))
    status = db.Column(db.String(20), default='offline')
    last_active = db.Column(db.DateTime)
    api_port = db.Column(db.Integer, default=5001)
    sessions = db.relationship('Session', backref='computer', lazy=True)

class Session(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    computer_id = db.Column(db.Integer, db.ForeignKey('computer.id'), nullable=False)
    login_time = db.Column(db.DateTime, nullable=False)
    logout_time = db.Column(db.DateTime)
    duration = db.Column(db.Integer)

class Advertisement(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    image_url = db.Column(db.String(500), nullable=False)
    link = db.Column(db.String(500), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.datetime.now)
    is_active = db.Column(db.Boolean, default=True)

class Background(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(255), nullable=False)
    url = db.Column(db.String(500), nullable=False)
    is_active = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.datetime.now)

class PendingEvent(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    event_type = db.Column(db.String(20), nullable=False)
    student_id = db.Column(db.String(20), nullable=False)
    computer_id = db.Column(db.Integer, nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.datetime.now)
    synced = db.Column(db.Boolean, default=False)
    event_uuid = db.Column(db.String(36), unique=True, nullable=False)

# ==== Khởi tạo DB ====
def init_db():
    with app.app_context():
        db.create_all()
        if Background.query.count() == 0:
            default_bg = Background(filename='default.jpg', url='static/images/default-bg.jpg', is_active=True)
            db.session.add(default_bg)
            db.session.commit()
        
        inspector = inspect(db.engine)
        
        if 'computer' in inspector.get_table_names():
            columns = [col['name'] for col in inspector.get_columns('computer')]
            if 'api_port' not in columns:
                try:
                    with db.engine.begin() as conn:
                        conn.execute(text('ALTER TABLE computer ADD COLUMN api_port INTEGER DEFAULT 5001'))
                except: pass
        
        if 'user' in inspector.get_table_names():
            columns = [col['name'] for col in inspector.get_columns('user')]
            if 'khoa_vien' not in columns:
                try:
                    with db.engine.begin() as conn:
                        conn.execute(text('ALTER TABLE "user" ADD COLUMN khoa_vien VARCHAR(100)'))
                except: pass
        
        migrate_ads_from_json_to_db()
        
        if Computer.query.count() == 0:
            for i in range(1, 67):
                computer = Computer(id=i, name=f'PC-{i:02d}', ip_address=None, mac_address=f'00:1A:2B:3C:{i:02d}:FF', status='offline', api_port=5001)
                db.session.add(computer)
            db.session.commit()
        
        admin = User.query.filter_by(student_id='admin').first()
        if not admin:
            admin = User(student_id='admin', name='Quản trị viên', class_name='ADMIN', khoa_vien='System', password_hash=generate_password_hash('admin123'), plain_password='admin123')
            db.session.add(admin)
            db.session.commit()

def migrate_ads_from_json_to_db():
    try:
        if Advertisement.query.count() > 0:
            return
        ads_path = resource_path('ads.json')
        if os.path.exists(ads_path):
            with open(ads_path, 'r', encoding='utf-8') as f:
                ads_data = json.load(f)
            for ad_data in ads_data:
                ad = Advertisement(image_url=ad_data.get('image_url', ''), link=ad_data.get('link', ''), is_active=True)
                db.session.add(ad)
            db.session.commit()
    except: pass

init_db()

# ==== Firebase ====
FIREBASE_URL = 'https://login-iuh-default-rtdb.firebaseio.com/'
def get_firebase_key_path():
    possible_paths = [resource_path('firebase_key.json'), os.path.join(os.path.dirname(__file__), "firebase_key.json"), "firebase_key.json"]
    for path in possible_paths:
        if os.path.exists(path):
            return path
    return None

FIREBASE_KEY_PATH = get_firebase_key_path()

def sync_user_to_firebase(user):
    if not FIREBASE_AVAILABLE or not firebase_admin._apps: return
    try:
        ref = firebase_db.reference(f'users/{user.student_id}')
        ref.set({'id': user.id, 'student_id': user.student_id, 'name': user.name, 'class_name': user.class_name, 'khoa_vien': user.khoa_vien, 'plain_password': user.plain_password or '', 'created_at': user.created_at.isoformat() if user.created_at else None})
    except: pass

def get_computer_name_from_ip(ip_address):
    try:
        parts = ip_address.strip().split('.')
        if len(parts) != 4: return None
        last_octet = int(parts[-1])
        if 1 <= last_octet <= 66: return f'PC-{last_octet:02d}'
        elif 101 <= last_octet <= 166: return f'PC-{last_octet-100:02d}'
        else: return f'PC-{(hash(ip_address) % 66 + 1):02d}'
    except: return None

# ==== API Endpoints ====
@app.route('/api/register_client', methods=['POST'])
def register_client():
    try:
        data = request.get_json()
        computer_id = data.get('computer_id')
        ip = data.get('ip')
        if not ip or ip.startswith('127.') or ip == '0.0.0.0' or ip == '::1':
            ip = request.remote_addr
        api_port = data.get('api_port', 5001)
        session_id = data.get('session_id')

        computer = Computer.query.filter_by(ip_address=ip).first()
        if not computer and computer_id:
            computer = Computer.query.get(computer_id)
        if not computer:
            computer_name = get_computer_name_from_ip(ip)
            if computer_name:
                computer = Computer.query.filter_by(name=computer_name).first()
                if not computer:
                    computer = Computer(name=computer_name, ip_address=ip, mac_address=f'00:1A:2B:3C:{computer_id:02d}:FF' if computer_id else '00:1A:2B:3C:00:00', status='online', api_port=api_port, last_active=datetime.datetime.now())
                    db.session.add(computer)
                    db.session.flush()

        if not computer:
            return jsonify({'success': False, 'message': 'Cannot identify computer'}), 400

        computer.ip_address = ip
        computer.api_port = api_port
        previous_last_active = computer.last_active
        computer.last_active = datetime.datetime.now()

        if session_id and isinstance(session_id, str) and session_id.startswith('offline-'):
            session_id = None

        if session_id:
            active_session = Session.query.filter_by(id=session_id, logout_time=None).first()
            if active_session and active_session.computer_id == computer.id:
                computer.status = 'Đang sử dụng'
            else:
                computer.status = 'online'
        else:
            active_session = Session.query.filter_by(computer_id=computer.id, logout_time=None).order_by(Session.login_time.desc()).first()
            if active_session:
                if previous_last_active and datetime.datetime.now() - previous_last_active > datetime.timedelta(minutes=15):
                    active_session.logout_time = previous_last_active
                    active_session.duration = int((active_session.logout_time - active_session.login_time).total_seconds())
                    computer.status = 'online'
                    active_session = None
                else:
                    computer.status = 'Đang sử dụng'
            else:
                computer.status = 'online'

        db.session.commit()
        client_apis[computer.id] = {'ip': ip, 'port': api_port}
        
        if active_session:
            user = User.query.get(active_session.user_id)
            return jsonify({'success': True, 'has_active_session': True, 'session_id': active_session.id, 'user_name': user.name if user else 'Unknown', 'user_id': user.id if user else None})
        return jsonify({'success': True, 'has_active_session': False})
    except Exception as e:
        logger.error(f"Register error: {e}")
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/heartbeat', methods=['POST'])
def heartbeat():
    try:
        data = request.get_json()
        computer_id = data.get('computer_id')
        computer = Computer.query.get(computer_id)
        if not computer and data.get('ip'):
            computer = Computer.query.filter_by(ip_address=data['ip']).first()
        if computer:
            computer.last_active = datetime.datetime.now()
            if data.get('ip'):
                computer.ip_address = data['ip']
            if data.get('api_port'):
                computer.api_port = data['api_port']
                client_apis[computer.id] = {'ip': computer.ip_address, 'port': computer.api_port}
            
            active_session = Session.query.filter_by(computer_id=computer.id, logout_time=None).first()
            computer.status = 'Đang sử dụng' if active_session else 'online'
            db.session.commit()
            return jsonify({'success': True, 'session_active': active_session is not None})
        return jsonify({'success': False}), 404
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/login', methods=['POST'])
def login():
    try:
        data = request.get_json()
        user = User.query.filter_by(student_id=data['student_id']).first()
        if not user:
            return jsonify({'success': False, 'message': 'Invalid credentials'}), 401
        password = data.get('password', '')
        if password:
            if not user.check_password(password):
                return jsonify({'success': False, 'message': 'Invalid credentials'}), 401
        
        computer = Computer.query.get(data['computer_id'])
        if not computer:
            return jsonify({'success': False, 'message': 'Computer not found'}), 404
        
        existing_session = Session.query.filter_by(computer_id=computer.id, logout_time=None).first()
        if existing_session:
            return jsonify({'success': False, 'message': 'Computer is in use'}), 409
        
        new_session = Session(user_id=user.id, computer_id=computer.id, login_time=datetime.datetime.now())
        computer.status = 'Đang sử dụng'
        computer.last_active = datetime.datetime.now()
        db.session.add(new_session)
        db.session.commit()
        
        return jsonify({'success': True, 'session_id': new_session.id, 'user_name': user.name})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/client_login', methods=['POST'])
def client_login_api():
    try:
        data = request.get_json()
        student_id = data.get('student_id')
        computer_id = data.get('computer_id')

        if not student_id or not computer_id:
            return jsonify({'success': False, 'message': 'Thiếu thông tin đăng nhập'}), 400

        user = User.query.filter_by(student_id=student_id).first()
        if not user:
            return jsonify({'success': False, 'message': 'Không tìm thấy user'}), 404

        computer = Computer.query.get(computer_id)
        if not computer:
            return jsonify({'success': False, 'message': 'Không tìm thấy máy tính'}), 404

        existing_session = Session.query.filter_by(computer_id=computer.id, logout_time=None).first()
        if existing_session:
            return jsonify({'success': False, 'message': 'Máy tính đang được sử dụng'}), 409

        new_session = Session(user_id=user.id, computer_id=computer.id, login_time=datetime.datetime.now())
        computer.status = 'Đang sử dụng'
        computer.last_active = datetime.datetime.now()
        db.session.add(new_session)
        db.session.commit()

        return jsonify({'success': True, 'session_id': new_session.id, 'user_name': user.name})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/sync_events', methods=['POST'])
def sync_events():
    try:
        data = request.get_json() or {}
        events = data.get('events', [])
        if not isinstance(events, list):
            return jsonify({'success': False, 'message': 'Dữ liệu không hợp lệ'}), 400

        synced_count = 0
        for event in events:
            event_type = event.get('event_type')
            student_id = str(event.get('student_id') or '').strip()
            computer_id = event.get('computer_id')
            timestamp = event.get('timestamp')

            if not event_type or not student_id or not computer_id:
                continue

            try:
                computer_id = int(computer_id)
            except (TypeError, ValueError):
                continue

            user = User.query.filter_by(student_id=student_id).first()
            computer = Computer.query.get(computer_id)
            if not user or not computer:
                continue

            event_time = datetime.datetime.now()
            if timestamp:
                try:
                    event_time = datetime.datetime.fromisoformat(timestamp)
                except Exception:
                    pass

            if event_type == 'login':
                existing_session = Session.query.filter_by(computer_id=computer.id, logout_time=None).first()
                if existing_session:
                    continue
                new_session = Session(user_id=user.id, computer_id=computer.id, login_time=event_time)
                computer.status = 'Đang sử dụng'
                computer.last_active = event_time
                db.session.add(new_session)
                synced_count += 1
            elif event_type == 'logout':
                active_session = Session.query.filter_by(computer_id=computer.id, logout_time=None).first()
                if not active_session:
                    continue
                active_session.logout_time = event_time
                active_session.duration = int((event_time - active_session.login_time).total_seconds())
                other_session = Session.query.filter_by(computer_id=computer.id, logout_time=None).filter(Session.id != active_session.id).first()
                computer.status = 'Đang sử dụng' if other_session else 'online'
                synced_count += 1

        db.session.commit()
        return jsonify({'success': True, 'synced_count': synced_count})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/qr_login', methods=['POST'])
def qr_login():
    try:
        data = request.get_json()
        student_id = data.get('student_id')
        user_id = data.get('user_id') or data.get('id')
        computer_id = data.get('computer_id')

        if student_id is not None:
            student_id = str(student_id).strip()
        if user_id is not None:
            user_id = str(user_id).strip()
        if isinstance(computer_id, str) and computer_id.isdigit():
            computer_id = int(computer_id)

        if not computer_id or (not student_id and not user_id):
            return jsonify({'success': False, 'message': 'Thiếu thông tin'}), 400

        user = None
        if student_id:
            user = User.query.filter_by(student_id=student_id).first()

        if not user and user_id:
            try:
                user_id_int = int(user_id)
                user = db.session.get(User, user_id_int)
            except (ValueError, TypeError):
                user = None

        if not user and student_id and student_id.isdigit():
            try:
                user = db.session.get(User, int(student_id))
            except (ValueError, TypeError):
                user = None

        if not user:
            return jsonify({'success': False, 'message': 'Không tìm thấy user'}), 404

        computer = db.session.get(Computer, computer_id)
        if not computer:
            return jsonify({'success': False, 'message': 'Không tìm thấy máy tính'}), 404

        existing_session = Session.query.filter_by(computer_id=computer.id, logout_time=None).first()
        if existing_session:
            return jsonify({'success': False, 'message': 'Máy tính đang được sử dụng bởi user khác'}), 409

        new_session = Session(user_id=user.id, computer_id=computer.id, login_time=datetime.datetime.now())
        computer.status = 'Đang sử dụng'
        computer.last_active = datetime.datetime.now()
        db.session.add(new_session)
        db.session.commit()

        return jsonify({'success': True, 'session_id': new_session.id, 'user_name': user.name})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/logout', methods=['POST'])
def logout():
    try:
        data = request.get_json()
        session_obj = Session.query.get(data['session_id'])
        if not session_obj or session_obj.logout_time:
            return jsonify({'success': False, 'message': 'Session not found'})
        
        session_obj.logout_time = datetime.datetime.now()
        session_obj.duration = int((session_obj.logout_time - session_obj.login_time).total_seconds())
        
        computer = Computer.query.get(session_obj.computer_id)
        if computer:
            other_session = Session.query.filter_by(computer_id=computer.id, logout_time=None).filter(Session.id != session_obj.id).first()
            computer.status = 'online' if not other_session else 'Đang sử dụng'
        
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/computers', methods=['GET'])
def get_computers():
    computers = Computer.query.order_by(Computer.id).all()
    return jsonify([{'id': c.id, 'name': c.name, 'status': c.status, 'ip_address': c.ip_address, 'last_active': c.last_active.isoformat() if c.last_active else None} for c in computers])

@app.route('/api/users', methods=['GET'])
def get_users():
    users = User.query.all()
    return jsonify([{'id': u.id, 'student_id': u.student_id, 'name': u.name, 'class_name': u.class_name, 'khoa_vien': u.khoa_vien, 'created_at': u.created_at.isoformat() if u.created_at else None} for u in users])

@app.route('/api/register', methods=['POST'])
def register():
    try:
        data = request.get_json()
        if User.query.filter_by(student_id=data['student_id']).first():
            return jsonify({'success': False, 'message': 'Student ID exists'}), 400
        if len(data['password']) < 8:
            return jsonify({'success': False, 'message': 'Password must be at least 8 characters'}), 400
        
        new_user = User(student_id=data['student_id'], name=data['name'], class_name=data['class_name'], khoa_vien=data.get('khoa_vien'), password_hash=generate_password_hash(data['password']), plain_password=data['password'])
        db.session.add(new_user)
        db.session.commit()
        sync_user_to_firebase(new_user)
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

# ==== REMOTE CONTROL APIs ====
@app.route('/api/remote/shutdown/<int:computer_id>', methods=['POST'])
def remote_shutdown(computer_id):
    try:
        if computer_id not in client_apis:
            computer = Computer.query.get(computer_id)
            if computer and computer.ip_address:
                client_apis[computer_id] = {'ip': computer.ip_address, 'port': computer.api_port or 5001}
            else:
                return jsonify({'success': False, 'message': 'Client not registered'}), 404
        
        client_info = client_apis[computer_id]
        url = f"http://{client_info['ip']}:{client_info['port']}/api/shutdown"
        
        for attempt in range(3):
            try:
                response = requests.post(url, timeout=5, json={})
                if response.status_code == 200:
                    computer = Computer.query.get(computer_id)
                    if computer:
                        computer.status = 'Tắt máy'
                        db.session.commit()
                    return jsonify({'success': True, 'message': 'Shutdown command sent'})
            except:
                time.sleep(1)
        return jsonify({'success': False, 'message': 'Failed after 3 attempts'}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/remote/restart/<int:computer_id>', methods=['POST'])
def remote_restart(computer_id):
    try:
        if computer_id not in client_apis:
            computer = Computer.query.get(computer_id)
            if computer and computer.ip_address:
                client_apis[computer_id] = {'ip': computer.ip_address, 'port': computer.api_port or 5001}
            else:
                return jsonify({'success': False, 'message': 'Client not registered'}), 404
        
        client_info = client_apis[computer_id]
        url = f"http://{client_info['ip']}:{client_info['port']}/api/restart"
        
        for attempt in range(3):
            try:
                response = requests.post(url, timeout=5, json={})
                if response.status_code == 200:
                    computer = Computer.query.get(computer_id)
                    if computer:
                        computer.status = 'Khởi động'
                        db.session.commit()
                    return jsonify({'success': True, 'message': 'Restart command sent'})
            except:
                time.sleep(1)
        return jsonify({'success': False, 'message': 'Failed after 3 attempts'}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/remote/screenshot/<int:computer_id>', methods=['GET'])
def remote_screenshot(computer_id):
    try:
        logger.info(f"Request remote screenshot for computer {computer_id}")
        if computer_id not in client_apis:
            computer = Computer.query.get(computer_id)
            if computer and computer.ip_address:
                client_apis[computer_id] = {'ip': computer.ip_address, 'port': computer.api_port or 5001}
            else:
                return jsonify({'success': False, 'message': 'Client not registered'}), 404
        
        client_info = client_apis[computer_id]
        url = f"http://{client_info['ip']}:{client_info['port']}/api/screenshot"
        logger.info(f"Forwarding screenshot request to client {client_info['ip']}:{client_info['port']}")
        response = requests.get(url, timeout=20)
        if response.status_code == 200 and response.content:
            logger.info(f"Received screenshot {len(response.content)} bytes from client {computer_id}")
            # Tạo response với cache headers
            img_response = send_file(io.BytesIO(response.content), mimetype='image/jpeg', as_attachment=False, download_name=f'screenshot_{computer_id}.jpg')
            # Set cache headers để browser không cache cũ
            img_response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
            img_response.headers['Pragma'] = 'no-cache'
            img_response.headers['Expires'] = '0'
            img_response.headers['Last-Modified'] = datetime.datetime.now().strftime('%a, %d %b %Y %H:%M:%S GMT')
            return img_response
        log_body = response.text if response is not None else 'no response'
        logger.warning(f"Screenshot request failed: status={response.status_code if response is not None else 'N/A'} body={log_body}")
        return jsonify({'success': False, 'message': 'No screenshot', 'detail': log_body}), 502
    except requests.exceptions.Timeout:
        logger.error(f"Timeout getting screenshot from {computer_id}")
        return jsonify({'success': False, 'message': 'Timeout'}), 504
    except requests.exceptions.ConnectionError as e:
        logger.error(f"Connection error for screenshot {computer_id}: {e}")
        return jsonify({'success': False, 'message': f'Cannot connect: {str(e)}'}), 503
    except Exception as e:
        logger.error(f"Screenshot error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/remote/logout/<int:computer_id>', methods=['POST'])
def remote_logout(computer_id):
    try:
        if computer_id not in client_apis:
            return jsonify({'success': False, 'message': 'Client not registered'}), 404
        
        client_info = client_apis[computer_id]
        url = f"http://{client_info['ip']}:{client_info['port']}/api/logout"
        response = requests.post(url, timeout=5)
        
        if response.status_code == 200:
            session_obj = Session.query.filter_by(computer_id=computer_id, logout_time=None).first()
            if session_obj:
                session_obj.logout_time = datetime.datetime.now()
                session_obj.duration = int((session_obj.logout_time - session_obj.login_time).total_seconds())
                computer = Computer.query.get(computer_id)
                if computer:
                    computer.status = 'online'
                db.session.commit()
            return jsonify({'success': True})
        return jsonify({'success': False}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/remote/status/<int:computer_id>', methods=['GET'])
def remote_status(computer_id):
    try:
        if computer_id not in client_apis:
            return jsonify({'online': False, 'message': 'Not registered'})
        
        client_info = client_apis[computer_id]
        url = f"http://{client_info['ip']}:{client_info['port']}/api/status"
        response = requests.get(url, timeout=3)
        
        if response.status_code == 200:
            return response.json()
        return jsonify({'online': False, 'message': f'HTTP {response.status_code}'})
    except Exception as e:
        return jsonify({'online': False, 'message': str(e)})

# ==== Debug APIs ====
@app.route('/api/debug/client_apis', methods=['GET'])
def debug_client_apis():
    result = {'registered': len(client_apis), 'clients': []}
    for cid, info in client_apis.items():
        computer = Computer.query.get(cid)
        result['clients'].append({
            'computer_id': cid,
            'name': computer.name if computer else f'PC-{cid}',
            'ip': info['ip'],
            'port': info['port'],
            'status': computer.status if computer else 'Unknown'
        })
    return jsonify(result)

# ==== Web Routes ====
@app.route('/')
def home():
    return redirect(url_for('admin_panel'))

@app.route('/admin')
def admin_panel():
    computers = Computer.query.order_by(Computer.id).all()
    users = User.query.order_by(User.student_id).all()
    
    # Xử lý partial request để làm mới danh sách máy
    if request.args.get('partial') == '1':
        return render_template('computer_management_partial.html', computers=computers, now=datetime.datetime.now())
    
    return render_template('admin.html', computers=computers, users=users, now=datetime.datetime.now())

@app.route('/client')
def client_login():
    return render_template('login.html')

# ==== Advertisement APIs ====
@app.route('/api/get_advertisement', methods=['GET'])
def get_advertisement():
    ads = Advertisement.query.filter_by(is_active=True).order_by(Advertisement.created_at.desc()).all()
    return jsonify([{'image_url': ad.image_url, 'link': ad.link} for ad in ads])

@app.route('/api/admin/advertisements', methods=['GET', 'POST'])
def admin_advertisements():
    if request.method == 'GET':
        ads = Advertisement.query.order_by(Advertisement.created_at.desc()).all()
        return jsonify([{'id': a.id, 'image_url': a.image_url, 'link': a.link, 'is_active': a.is_active, 'created_at': a.created_at.isoformat() if a.created_at else None} for a in ads])
    
    try:
        if 'image' not in request.files:
            return jsonify({'success': False, 'message': 'No image'}), 400
        file = request.files['image']
        link = request.form.get('link')
        if not link:
            return jsonify({'success': False, 'message': 'Link required'}), 400
        
        filename = secure_filename(f"{int(time.time())}_{file.filename}")
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        ad = Advertisement(image_url=f"static/uploads/ads/{filename}", link=link, is_active=True)
        db.session.add(ad)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/admin/advertisements/<int:ad_id>', methods=['DELETE'])
def delete_advertisement(ad_id):
    try:
        ad = Advertisement.query.get_or_404(ad_id)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], ad.image_url.replace('static/uploads/ads/', ''))
        if os.path.exists(filepath):
            os.remove(filepath)
        db.session.delete(ad)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

# ==== Background APIs ====
@app.route('/api/get_backgrounds', methods=['GET'])
def get_backgrounds():
    backgrounds = Background.query.all()
    return jsonify({'backgrounds': [{'filename': b.filename, 'url': url_for('static', filename=f'uploads/backgrounds/{b.filename}', _external=True), 'is_active': b.is_active} for b in backgrounds]})

@app.route('/api/get_active_background', methods=['GET'])
def get_active_background():
    bg = Background.query.filter_by(is_active=True).first()
    if bg:
        # Trả về URL đầy đủ (absolute) để client có thể tải ảnh trực tiếp
        absolute_url = url_for('static', filename=f'uploads/backgrounds/{bg.filename}', _external=True)
        return jsonify({'success': True, 'url': absolute_url})
    return jsonify({'success': False}), 404

@app.route('/api/upload_background', methods=['POST'])
def upload_background():
    if 'background' not in request.files:
        return jsonify({'success': False, 'message': 'No file'}), 400
    file = request.files['background']
    if file and allowed_file(file.filename):
        filename = secure_filename(f"{int(time.time())}_{file.filename}")
        filepath = os.path.join(app.config['BACKGROUND_FOLDER'], filename)
        file.save(filepath)
        new_bg = Background(filename=filename, url=f"static/uploads/backgrounds/{filename}", is_active=False)
        db.session.add(new_bg)
        db.session.commit()
        return jsonify({'success': True, 'url': f"static/uploads/backgrounds/{filename}"})
    return jsonify({'success': False, 'message': 'Invalid file'}), 400

@app.route('/api/set_active_background', methods=['POST'])
def set_active_background():
    data = request.get_json()
    filename = data.get('filename')
    Background.query.update({'is_active': False})
    bg = Background.query.filter_by(filename=filename).first()
    if bg:
        bg.is_active = True
        db.session.commit()
        return jsonify({'success': True})
    return jsonify({'success': False}), 404

# ==== Background Tasks ====
def check_computers_status():
    while True:
        try:
            now = datetime.datetime.now()
            timeout = datetime.timedelta(seconds=app.config['HEARTBEAT_TIMEOUT'])
            
            with app.app_context():
                computers = Computer.query.all()
                for computer in computers:
                    if computer.last_active is None:
                        continue
                    
                    active_session = Session.query.filter_by(computer_id=computer.id, logout_time=None).first()
                    time_since = now - computer.last_active
                    
                    if time_since > timeout * 3:
                        computer.status = 'offline'
                        if active_session:
                            active_session.logout_time = now
                            active_session.duration = int((now - active_session.login_time).total_seconds())
                    elif time_since > timeout:
                        computer.status = 'Mất kết nối' if active_session else 'offline'
                    else:
                        computer.status = 'Đang sử dụng' if active_session else 'online'
                
                db.session.commit()
            time.sleep(app.config['HEARTBEAT_INTERVAL'])
        except Exception as e:
            logger.error(f"Status check error: {e}")
            time.sleep(5)

def start_background_tasks():
    status_thread = Thread(target=check_computers_status, daemon=True)
    status_thread.start()

start_background_tasks()

ALL_KHOA_VIEN = [
    "Khoa Công nghệ Cơ khí", "Khoa Công nghệ Thông tin", "Khoa Công nghệ Điện",
    "Khoa Công nghệ Điện tử", "Khoa Công nghệ Động lực", "Khoa Công nghệ Nhiệt - Lạnh",
    "Khoa Công nghệ May - Thời trang", "Khoa Công nghệ Hóa học", "Khoa Ngoại ngữ",
    "Khoa Quản trị Kinh doanh", "Khoa Thương mại - Du lịch", "Khoa Kỹ thuật Xây dựng",
    "Khoa Luật", "Viện Tài chính - Kế toán", "Viện Công nghệ Sinh học và Thực phẩm",
    "Viện Khoa học Công nghệ và Quản lý Môi trường", "Khoa Khoa học Cơ bản"
]

@app.route('/api/report/khoa_vien_range', methods=['GET'])
def report_khoa_vien_range():
    from_date = request.args.get('from')
    to_date = request.args.get('to')
    try:
        from_dt = datetime.datetime.strptime(from_date, '%Y-%m-%d')
        to_dt = datetime.datetime.strptime(to_date, '%Y-%m-%d') + datetime.timedelta(days=1)
        
        results = db.session.query(User.khoa_vien, db.func.count(Session.id)).join(Session, Session.user_id == User.id).filter(Session.login_time >= from_dt, Session.login_time < to_dt).group_by(User.khoa_vien).all()
        result_dict = {r[0]: r[1] for r in results}
        
        report = [{'khoa_vien': khoa, 'so_luot': result_dict.get(khoa, 0)} for khoa in ALL_KHOA_VIEN]
        return jsonify(report)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        return "127.0.0.1"

def open_browser():
    time.sleep(2)
    ip = get_local_ip()
    webbrowser.open_new(f'http://{ip}:5000')

# ============ THÊM CÁC API NÀY VÀO app.py (thêm trước if __name__ == '__main__') ============

@app.route('/api/remote/force_register/<int:computer_id>', methods=['POST'])
def force_register_client(computer_id):
    """Force server to register a client by IP from database"""
    try:
        computer = Computer.query.get(computer_id)
        if not computer:
            return jsonify({'success': False, 'message': 'Computer not found'}), 404
        
        if not computer.ip_address:
            return jsonify({'success': False, 'message': 'Computer has no IP address'}), 404
        
        # Force register
        client_apis[computer.id] = {
            'ip': computer.ip_address,
            'port': computer.api_port or 5001
        }
        
        return jsonify({
            'success': True,
            'message': f'Registered {computer.name} with IP {computer.ip_address}',
            'computer': {
                'id': computer.id,
                'name': computer.name,
                'ip': computer.ip_address,
                'port': computer.api_port
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/remote/shutdown/<int:computer_id>', methods=['POST'])
def remote_shutdown_fixed(computer_id):
    """Improved shutdown with better error handling"""
    try:
        logger.info(f"🔌 SHUTDOWN request for computer {computer_id}")
        
        # Try to get from client_apis first
        client_info = client_apis.get(computer_id)
        
        # If not found, try to get from database
        if not client_info:
            computer = Computer.query.get(computer_id)
            if computer and computer.ip_address:
                client_info = {
                    'ip': computer.ip_address,
                    'port': computer.api_port or 5001
                }
                # Cache for next time
                client_apis[computer_id] = client_info
                logger.info(f"📝 Cached client {computer_id} from DB: {client_info}")
            else:
                return jsonify({
                    'success': False,
                    'message': f'Computer {computer_id} not registered',
                    'debug': {
                        'computer_exists': computer is not None,
                        'computer_ip': computer.ip_address if computer else None,
                        'registered_clients': list(client_apis.keys())
                    }
                }), 404
        
        # Send shutdown command with retry
        url = f"http://{client_info['ip']}:{client_info['port']}/api/shutdown"
        logger.info(f"🔌 Sending shutdown to {url}")
        
        for attempt in range(3):
            try:
                response = requests.post(url, timeout=5, json={})
                if response.status_code == 200:
                    computer = Computer.query.get(computer_id)
                    if computer:
                        computer.status = 'Tắt máy'
                        db.session.commit()
                    logger.info(f"✅ Shutdown successful for {computer_id}")
                    return jsonify({'success': True, 'message': 'Shutdown command sent'})
                else:
                    logger.warning(f"Attempt {attempt+1}: HTTP {response.status_code}")
            except requests.exceptions.Timeout:
                logger.warning(f"Attempt {attempt+1}: Timeout")
            except requests.exceptions.ConnectionError as e:
                logger.warning(f"Attempt {attempt+1}: Connection error: {e}")
            time.sleep(1)
        
        return jsonify({
            'success': False,
            'message': 'Failed to send shutdown command after 3 attempts',
            'url': url
        }), 500
        
    except Exception as e:
        logger.error(f"Shutdown error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/remote/restart/<int:computer_id>', methods=['POST'])
def remote_restart_fixed(computer_id):
    """Improved restart with better error handling"""
    try:
        logger.info(f"🔄 RESTART request for computer {computer_id}")
        
        client_info = client_apis.get(computer_id)
        if not client_info:
            computer = Computer.query.get(computer_id)
            if computer and computer.ip_address:
                client_info = {'ip': computer.ip_address, 'port': computer.api_port or 5001}
                client_apis[computer_id] = client_info
            else:
                return jsonify({'success': False, 'message': f'Computer {computer_id} not registered'}), 404
        
        url = f"http://{client_info['ip']}:{client_info['port']}/api/restart"
        logger.info(f"🔄 Sending restart to {url}")
        
        for attempt in range(3):
            try:
                response = requests.post(url, timeout=5, json={})
                if response.status_code == 200:
                    computer = Computer.query.get(computer_id)
                    if computer:
                        computer.status = 'Khởi động'
                        db.session.commit()
                    logger.info(f"✅ Restart successful for {computer_id}")
                    return jsonify({'success': True, 'message': 'Restart command sent'})
            except:
                time.sleep(1)
        
        return jsonify({'success': False, 'message': 'Failed after 3 attempts'}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/remote/screenshot/<int:computer_id>', methods=['GET'])
def remote_screenshot_fixed(computer_id):
    """Improved screenshot with better timeout"""
    try:
        logger.info(f"📸 SCREENSHOT request for computer {computer_id}")
        
        client_info = client_apis.get(computer_id)
        if not client_info:
            computer = Computer.query.get(computer_id)
            if computer and computer.ip_address:
                client_info = {'ip': computer.ip_address, 'port': computer.api_port or 5001}
                client_apis[computer_id] = client_info
            else:
                return jsonify({'success': False, 'message': 'Client not registered'}), 404
        
        url = f"http://{client_info['ip']}:{client_info['port']}/api/screenshot"
        
        # Shorter timeout for faster response
        response = requests.get(url, timeout=5)
        
        if response.status_code == 200 and response.content:
            logger.info(f"✅ Screenshot received: {len(response.content)} bytes")
            img_response = send_file(
                io.BytesIO(response.content),
                mimetype='image/jpeg',
                as_attachment=False,
                download_name=f'screenshot_{computer_id}.jpg'
            )
            # Set cache headers để browser không cache cũ
            img_response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
            img_response.headers['Pragma'] = 'no-cache'
            img_response.headers['Expires'] = '0'
            img_response.headers['Last-Modified'] = datetime.datetime.now().strftime('%a, %d %b %Y %H:%M:%S GMT')
            return img_response
        else:
            return jsonify({'success': False, 'message': 'No screenshot data'}), 502
            
    except requests.exceptions.Timeout:
        logger.error(f"Timeout getting screenshot from {computer_id}")
        return jsonify({'success': False, 'message': 'Timeout - Client did not respond'}), 504
    except requests.exceptions.ConnectionError as e:
        logger.error(f"Connection error: {e}")
        return jsonify({'success': False, 'message': f'Cannot connect: {str(e)}'}), 503
    except Exception as e:
        logger.error(f"Screenshot error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/remote/status/<int:computer_id>', methods=['GET'])
def remote_status_fixed(computer_id):
    """Check if client is online"""
    try:
        client_info = client_apis.get(computer_id)
        if not client_info:
            computer = Computer.query.get(computer_id)
            if computer and computer.ip_address:
                client_info = {'ip': computer.ip_address, 'port': computer.api_port or 5001}
            else:
                return jsonify({'online': False, 'message': 'Not registered'})
        
        url = f"http://{client_info['ip']}:{client_info['port']}/api/status"
        response = requests.get(url, timeout=3)
        
        if response.status_code == 200:
            return response.json()
        return jsonify({'online': False, 'message': f'HTTP {response.status_code}'})
    except Exception as e:
        return jsonify({'online': False, 'message': str(e)})


@app.route('/api/remote/broadcast_shutdown', methods=['POST'])
def broadcast_shutdown():
    """Shutdown all online clients"""
    results = {'success': [], 'failed': []}
    
    for computer_id, client_info in list(client_apis.items()):
        try:
            computer = Computer.query.get(computer_id)
            url = f"http://{client_info['ip']}:{client_info['port']}/api/shutdown"
            response = requests.post(url, timeout=3, json={})
            
            if response.status_code == 200:
                results['success'].append(computer_id)
                if computer:
                    computer.status = 'Tắt máy'
            else:
                results['failed'].append(computer_id)
        except Exception as e:
            results['failed'].append(computer_id)
            logger.error(f"Broadcast shutdown failed for {computer_id}: {e}")
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'results': results,
        'total': len(client_apis),
        'success_count': len(results['success']),
        'failed_count': len(results['failed'])
    })


@app.route('/api/debug/check_connection/<int:computer_id>', methods=['GET'])
def debug_check_connection(computer_id):
    """Debug endpoint to test connection to client"""
    try:
        computer = Computer.query.get(computer_id)
        if not computer:
            return jsonify({'success': False, 'message': 'Computer not found'}), 404
        
        result = {
            'computer_id': computer_id,
            'computer_name': computer.name,
            'computer_ip': computer.ip_address,
            'computer_port': computer.api_port,
            'computer_status': computer.status,
            'in_client_apis': computer_id in client_apis,
            'client_apis_info': client_apis.get(computer_id),
            'ping_test': None,
            'port_test': None,
            'api_test': None
        }
        
        # Test ping
        import subprocess
        try:
            ping = subprocess.run(['ping', '-n', '1', '-w', '1000', computer.ip_address], 
                                 capture_output=True, text=True, timeout=3)
            result['ping_test'] = ping.returncode == 0
        except:
            result['ping_test'] = False
        
        # Test port
        if computer.ip_address:
            try:
                sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                sock.settimeout(2)
                port_result = sock.connect_ex((computer.ip_address, computer.api_port or 5001))
                sock.close()
                result['port_test'] = port_result == 0
            except:
                result['port_test'] = False
        
        # Test API
        if computer.ip_address and result.get('port_test'):
            try:
                url = f"http://{computer.ip_address}:{computer.api_port or 5001}/api/status"
                response = requests.get(url, timeout=3)
                result['api_test'] = {
                    'success': response.status_code == 200,
                    'status_code': response.status_code,
                    'response': response.json() if response.status_code == 200 else None
                }
            except Exception as e:
                result['api_test'] = {'success': False, 'error': str(e)}
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
# ============ THÊM CÁC ROUTE SAU VÀO app.py ============

@app.route('/admin/shutdown_all', methods=['POST'])
def shutdown_all():
    """Tắt tất cả máy online - gọi trực tiếp tới tất cả client đã đăng ký"""
    errors = []
    success_count = 0
    
    # Tắt tất cả client_apis
    for computer_id, client_info in list(client_apis.items()):
        try:
            computer = Computer.query.get(computer_id)
            if not computer:
                continue
                
            client_url = f"http://{client_info['ip']}:{client_info['port']}/api/shutdown"
            logger.info(f"🔌 [SHUTDOWN_ALL] Gửi shutdown đến {computer.name}")
            
            response = requests.post(client_url, timeout=3, json={})
            if response.status_code == 200:
                computer.status = 'Tắt máy'
                db.session.commit()
                success_count += 1
            else:
                errors.append(computer.name)
        except requests.exceptions.RequestException:
            if computer:
                errors.append(computer.name)
        except Exception:
            if computer:
                errors.append(computer.name)
    
    if success_count > 0:
        flash(f"✓ Đã gửi lệnh tắt thành công đến {success_count} máy", 'success')
    if errors:
        flash(f"❌ Không thể tắt: {', '.join(errors)}", 'error')
    if success_count == 0 and not errors:
        flash("⚠️ Không có máy nào online", 'warning')
    
    return redirect(url_for('admin_panel'))


@app.route('/admin/restart_all', methods=['POST'])
def restart_all():
    """Khởi động lại tất cả máy online - gọi trực tiếp tới tất cả client đã đăng ký"""
    errors = []
    success_count = 0
    
    # Khởi động lại tất cả client_apis
    for computer_id, client_info in list(client_apis.items()):
        try:
            computer = Computer.query.get(computer_id)
            if not computer:
                continue
                
            client_url = f"http://{client_info['ip']}:{client_info['port']}/api/restart"
            logger.info(f"🔄 [RESTART_ALL] Gửi restart đến {computer.name}")
            
            response = requests.post(client_url, timeout=3, json={})
            if response.status_code == 200:
                computer.status = 'Khởi động'
                db.session.commit()
                success_count += 1
            else:
                errors.append(computer.name)
        except requests.exceptions.RequestException:
            if computer:
                errors.append(computer.name)
        except Exception:
            if computer:
                errors.append(computer.name)
    
    if success_count > 0:
        flash(f"✓ Đã gửi lệnh khởi động lại thành công đến {success_count} máy", 'success')
    if errors:
        flash(f"❌ Không thể khởi động lại: {', '.join(errors)}", 'error')
    if success_count == 0 and not errors:
        flash("⚠️ Không có máy nào online", 'warning')
    
    return redirect(url_for('admin_panel'))


@app.route('/admin/shutdown_online', methods=['POST'])
def shutdown_online():
    """Tắt tất cả máy online - alias cho shutdown_all"""
    return shutdown_all()


@app.route('/admin/restart_online', methods=['POST'])
def restart_online():
    """Khởi động lại tất cả máy online - alias cho restart_all"""
    return restart_all()


@app.route('/admin/shutdown/<int:computer_id>')
def admin_shutdown(computer_id):
    """Tắt máy từ admin panel"""
    try:
        if computer_id not in client_apis:
            computer = Computer.query.get(computer_id)
            if computer and computer.ip_address:
                client_apis[computer_id] = {'ip': computer.ip_address, 'port': computer.api_port or 5001}
            else:
                flash(f'Máy {computer_id} không online hoặc chưa đăng ký', 'error')
                return redirect(url_for('admin_panel'))
        
        client_info = client_apis[computer_id]
        client_url = f"http://{client_info['ip']}:{client_info['port']}/api/shutdown"
        
        response = requests.post(client_url, timeout=5, json={})
        
        if response.status_code == 200:
            computer = Computer.query.get(computer_id)
            if computer:
                computer.status = 'Tắt máy'
                db.session.commit()
            flash(f'✓ Đã gửi lệnh tắt máy thành công', 'success')
        else:
            flash('❌ Không thể gửi lệnh tắt máy - client không phản hồi', 'error')
    except requests.exceptions.RequestException as e:
        flash(f'❌ Không thể kết nối đến client: {str(e)}', 'error')
    except Exception as e:
        flash(f'❌ Lỗi: {str(e)}', 'error')
    
    return redirect(url_for('admin_panel'))


@app.route('/admin/restart/<int:computer_id>')
def admin_restart(computer_id):
    """Khởi động lại máy từ admin panel"""
    try:
        if computer_id not in client_apis:
            computer = Computer.query.get(computer_id)
            if computer and computer.ip_address:
                client_apis[computer_id] = {'ip': computer.ip_address, 'port': computer.api_port or 5001}
            else:
                flash(f'Máy {computer_id} không online hoặc chưa đăng ký', 'error')
                return redirect(url_for('admin_panel'))
        
        client_info = client_apis[computer_id]
        client_url = f"http://{client_info['ip']}:{client_info['port']}/api/restart"
        
        response = requests.post(client_url, timeout=5, json={})
        
        if response.status_code == 200:
            computer = Computer.query.get(computer_id)
            if computer:
                computer.status = 'Khởi động'
                db.session.commit()
            flash(f'✓ Đã gửi lệnh khởi động lại thành công', 'success')
        else:
            flash('❌ Không thể gửi lệnh khởi động lại - client không phản hồi', 'error')
    except requests.exceptions.RequestException as e:
        flash(f'❌ Không thể kết nối đến client: {str(e)}', 'error')
    except Exception as e:
        flash(f'❌ Lỗi: {str(e)}', 'error')
    
    return redirect(url_for('admin_panel'))


@app.route('/admin/logout/<int:computer_id>')
def admin_logout(computer_id):
    """Đăng xuất user trên client từ admin panel"""
    try:
        if computer_id not in client_apis:
            computer = Computer.query.get(computer_id)
            if computer and computer.ip_address:
                client_apis[computer_id] = {'ip': computer.ip_address, 'port': computer.api_port or 5001}
            else:
                flash('Client chưa đăng ký hoặc offline', 'error')
                return redirect(url_for('admin_panel'))
        
        client_info = client_apis[computer_id]
        client_url = f"http://{client_info['ip']}:{client_info['port']}/api/logout"
        
        response = requests.post(client_url, timeout=5)
        
        if response.status_code == 200:
            active_session = Session.query.filter_by(computer_id=computer_id, logout_time=None).first()
            if active_session:
                now = datetime.datetime.now()
                active_session.logout_time = now
                active_session.duration = int((now - active_session.login_time).total_seconds())
                computer = Computer.query.get(computer_id)
                if computer:
                    computer.status = 'online'
                db.session.commit()
            flash('Đã gửi lệnh đăng xuất tới client!', 'success')
        else:
            flash('Không thể gửi lệnh đăng xuất tới client', 'error')
    except Exception as e:
        flash(f'Lỗi khi gửi lệnh đăng xuất: {str(e)}', 'error')
    
    return redirect(url_for('admin_panel'))


@app.route('/admin/remote_login/<int:computer_id>')
def admin_remote_login(computer_id):
    """Đăng nhập từ xa vào client từ admin panel"""
    try:
        if computer_id not in client_apis:
            computer = Computer.query.get(computer_id)
            if computer and computer.ip_address:
                client_apis[computer_id] = {'ip': computer.ip_address, 'port': computer.api_port or 5001}
            else:
                flash('Client chưa đăng ký hoặc offline', 'error')
                return redirect(url_for('admin_panel'))
        
        client_info = client_apis[computer_id]
        client_url = f"http://{client_info['ip']}:{client_info['port']}/api/remote_login"
        
        response = requests.post(client_url, timeout=5)
        
        if response.status_code == 200:
            flash('Đã gửi lệnh đăng nhập từ xa tới client!', 'success')
        else:
            flash('Không thể gửi lệnh đăng nhập từ xa tới client', 'error')
    except Exception as e:
        flash(f'Lỗi khi gửi lệnh đăng nhập từ xa: {str(e)}', 'error')
    
    return redirect(url_for('admin_panel'))


@app.route('/admin/screenshot/<int:computer_id>')
def admin_screenshot(computer_id):
    """Xem screenshot từ admin panel"""
    try:
        if computer_id not in client_apis:
            computer = Computer.query.get(computer_id)
            if computer and computer.ip_address:
                client_apis[computer_id] = {'ip': computer.ip_address, 'port': computer.api_port or 5001}
            else:
                flash('Client chưa đăng ký hoặc offline', 'error')
                return redirect(url_for('admin_panel'))
        
        client_info = client_apis[computer_id]
        screenshot_url = f"http://{client_info['ip']}:{client_info['port']}/api/screenshot"
        
        response = requests.get(screenshot_url, timeout=15)
        
        if response.status_code == 200 and response.content:
            return send_file(
                io.BytesIO(response.content),
                mimetype='image/jpeg',
                as_attachment=False,
                download_name=f'screenshot_{computer_id}.jpg'
            )
        else:
            flash('Không thể lấy screenshot từ client', 'error')
            return redirect(url_for('admin_panel'))
            
    except requests.exceptions.Timeout:
        flash('Timeout - client không phản hồi', 'error')
    except Exception as e:
        flash(f'Lỗi: {str(e)}', 'error')
    
    return redirect(url_for('admin_panel'))


# Thêm API endpoint cho screenshot (GET)
@app.route('/get_screenshot/<int:computer_id>')
def get_screenshot(computer_id):
    """Lấy screenshot từ client (API endpoint)"""
    return admin_screenshot(computer_id)


# Thêm API endpoint cho upload screenshot
@app.route('/upload_screenshot', methods=['POST'])
def upload_screenshot():
    """Client upload screenshot lên server"""
    try:
        computer_id = request.form.get('computer_id')
        file = request.files.get('screenshot')
        
        if not computer_id or not file:
            return jsonify({'success': False, 'message': 'Missing data'}), 400
        
        # Lưu screenshot vào memory hoặc disk
        if not hasattr(app, 'latest_screenshots'):
            app.latest_screenshots = {}
        
        app.latest_screenshots[computer_id] = file.read()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# ============ THÊM CÁC ROUTE IMPORT/EXPORT USERS ============

@app.route('/admin/import_users', methods=['GET', 'POST'])
def import_users():
    """Import users from Excel file"""
    if request.method == 'POST':
        file = request.files.get('file')
        if not file:
            flash('Vui lòng chọn file Excel!', 'error')
            return redirect(url_for('admin_panel'))
        
        try:
            # Đọc file Excel
            df = pd.read_excel(file, dtype={'student_id': str})
            
            # Các cột bắt buộc
            required_columns = ['student_id', 'name', 'class_name', 'khoa_vien', 'password']
            
            # Kiểm tra các cột cần thiết
            for col in required_columns:
                if col not in df.columns:
                    flash(f'Thiếu cột "{col}" trong file Excel!', 'error')
                    return redirect(url_for('admin_panel'))
            
            success_count = 0
            error_count = 0
            errors = []
            
            for idx, row in df.iterrows():
                try:
                    student_id = str(row['student_id']).strip()
                    password = str(row['password']).strip()
                    name = str(row['name']).strip()
                    class_name = str(row['class_name']).strip()
                    khoa_vien = str(row['khoa_vien']).strip() if pd.notna(row['khoa_vien']) else ''
                    
                    # Kiểm tra user đã tồn tại chưa
                    if User.query.filter_by(student_id=student_id).first():
                        error_count += 1
                        errors.append(f"Dòng {idx+2}: Mã sinh viên {student_id} đã tồn tại")
                        continue
                    
                    # Kiểm tra mật khẩu
                    if len(password) < 8:
                        error_count += 1
                        errors.append(f"Dòng {idx+2}: Mật khẩu của {student_id} phải có ít nhất 8 ký tự")
                        continue
                    
                    # Tạo user mới
                    new_user = User(
                        student_id=student_id,
                        name=name,
                        class_name=class_name,
                        khoa_vien=khoa_vien,
                        password_hash=generate_password_hash(password),
                        plain_password=password
                    )
                    db.session.add(new_user)
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    errors.append(f"Dòng {idx+2}: {str(e)}")
            
            db.session.commit()
            
            # Hiển thị kết quả
            if success_count > 0:
                flash(f'✓ Import thành công {success_count} người dùng!', 'success')
            if error_count > 0:
                flash(f'✗ Có {error_count} lỗi: {" | ".join(errors[:5])}{"..." if len(errors) > 5 else ""}', 'error')
                
        except Exception as e:
            db.session.rollback()
            flash(f'Lỗi khi đọc file: {str(e)}', 'error')
        
        return redirect(url_for('admin_panel'))
    
    # GET request - hiển thị form import
    return render_template('import_users.html')


@app.route('/admin/export_users_excel')
def export_users_excel():
    """Export all users to Excel file"""
    if not OPENPYXL_AVAILABLE:
        return "Excel export not available - openpyxl not installed", 503
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        
        # Header
        headers = ["student_id", "name", "class_name", "khoa_vien", "password"]
        ws.append(headers)
        
        # Format header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        users = User.query.all()
        for user in users:
            ws.append([
                user.student_id,
                user.name,
                user.class_name or '',
                user.khoa_vien or '',
                ''  # Không export password để bảo mật
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Export file
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            download_name=f"users_export_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Lỗi export: {str(e)}', 'error')
        return redirect(url_for('admin_panel'))


@app.route('/admin/export_users_template')
def export_users_template():
    """Download template Excel file for importing users"""
    if not OPENPYXL_AVAILABLE:
        return "Excel export not available - openpyxl not installed", 503
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "UsersTemplate"
        
        # Header
        headers = ["student_id", "name", "class_name", "khoa_vien", "password"]
        ws.append(headers)
        
        # Format header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Example data
        example_data = [
            ["20123456", "Nguyễn Văn A", "DHTH17A", "Khoa Công nghệ Thông tin", "12345678"],
            ["20123457", "Trần Thị B", "DHTH17B", "Khoa Công nghệ Thông tin", "12345678"],
            ["20123458", "Lê Văn C", "DHKT18A", "Khoa Công nghệ Cơ khí", "12345678"],
        ]
        
        for row_data in example_data:
            ws.append(row_data)
        
        # Add instruction comment
        ws['G1'] = "Hướng dẫn:"
        ws['G2'] = "- student_id: Mã số sinh viên (bắt buộc, duy nhất)"
        ws['G3'] = "- name: Họ tên đầy đủ (bắt buộc)"
        ws['G4'] = "- class_name: Tên lớp (bắt buộc)"
        ws['G5'] = "- khoa_vien: Tên khoa/viện (bắt buộc)"
        ws['G6'] = "- password: Mật khẩu (bắt buộc, tối thiểu 8 ký tự)"
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Export file
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            download_name="users_import_template.xlsx",
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Lỗi tạo template: {str(e)}', 'error')
        return redirect(url_for('admin_panel'))


# Thêm route để xóa user
@app.route('/admin/delete_user/<int:user_id>', methods=['POST'])
def delete_user(user_id):
    """Delete a user"""
    try:
        user = User.query.get(user_id)
        if not user:
            flash('Không tìm thấy người dùng!', 'error')
            return redirect(url_for('admin_panel'))
        
        # Xóa sessions trước
        Session.query.filter_by(user_id=user.id).delete()
        
        # Xóa user
        db.session.delete(user)
        db.session.commit()
        
        # Xóa từ Firebase nếu có
        if FIREBASE_AVAILABLE and firebase_admin._apps:
            try:
                ref = firebase_db.reference(f'users/{user.student_id}')
                ref.delete()
            except:
                pass
        
        flash(f'✓ Đã xóa người dùng {user.student_id} - {user.name}', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Lỗi khi xóa: {str(e)}', 'error')
    
    return redirect(url_for('admin_panel'))


# Thêm route để cập nhật user
@app.route('/admin/update_user/<int:user_id>', methods=['POST'])
def update_user(user_id):
    """Update user information"""
    try:
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404
        
        data = request.get_json()
        
        if 'student_id' in data:
            user.student_id = data['student_id']
        if 'name' in data:
            user.name = data['name']
        if 'class_name' in data:
            user.class_name = data['class_name']
        if 'khoa_vien' in data:
            user.khoa_vien = data['khoa_vien']
        if 'password' in data and data['password']:
            if len(data['password']) >= 8:
                user.set_password(data['password'])
                user.plain_password = data['password']
            else:
                return jsonify({'success': False, 'message': 'Password must be at least 8 characters'}), 400
        
        db.session.commit()
        
        # Sync to Firebase
        sync_user_to_firebase(user)
        
        return jsonify({'success': True, 'message': 'User updated successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


# API để lấy danh sách users (có thể đã có, nhưng đảm bảo tồn tại)
@app.route('/api/users/list', methods=['GET'])
def api_users_list():
    """Get list of users for admin panel"""
    search = request.args.get('search', '').strip()
    query = User.query
    
    if search:
        like = f"%{search}%"
        query = query.filter(
            or_(
                User.student_id.ilike(like),
                User.name.ilike(like),
                User.class_name.ilike(like),
                User.khoa_vien.ilike(like)
            )
        )
    
    users = query.order_by(User.id.desc()).all()
    
    return jsonify({
        'success': True,
        'users': [{
            'id': u.id,
            'student_id': u.student_id,
            'name': u.name,
            'class_name': u.class_name,
            'khoa_vien': u.khoa_vien or '',
            'created_at': u.created_at.isoformat() if u.created_at else None
        } for u in users],
        'total': len(users)
    })


@app.route('/admin/users', methods=['GET'])
def admin_users():
    """Get list of users for admin panel user management page"""
    search = request.args.get('search', '').strip()
    query = User.query
    
    if search:
        like = f"%{search}%"
        query = query.filter(
            or_(
                User.student_id.ilike(like),
                User.name.ilike(like),
                User.class_name.ilike(like),
                User.khoa_vien.ilike(like)
            )
        )
    
    users = query.order_by(User.id.desc()).all()
    
    return jsonify({
        'users': [{
            'id': u.id,
            'student_id': u.student_id,
            'name': u.name,
            'class_name': u.class_name,
            'khoa_vien': u.khoa_vien or '',
            'created_at': u.created_at.isoformat() if u.created_at else None
        } for u in users],
        'total': len(users)
    })

@app.route('/admin/add_user', methods=['POST'])
def add_user():
    """Add a new user from admin panel"""
    try:
        data = request.get_json()
        student_id = data.get('student_id', '').strip()
        name = data.get('name', '').strip()
        class_name = data.get('class_name', '').strip()
        khoa_vien = data.get('khoa_vien', '').strip()
        password = data.get('password', '').strip()
        
        if not student_id or not name or not password:
            return jsonify({'success': False, 'message': 'Mã SV, Họ tên và Mật khẩu là bắt buộc'}), 400
        
        if User.query.filter_by(student_id=student_id).first():
            return jsonify({'success': False, 'message': 'Mã sinh viên đã tồn tại'}), 400
        
        if len(password) < 8:
            return jsonify({'success': False, 'message': 'Mật khẩu phải có ít nhất 8 ký tự'}), 400
        
        user = User(student_id=student_id, name=name, class_name=class_name, khoa_vien=khoa_vien)
        user.set_password(password)
        user.plain_password = password
        
        db.session.add(user)
        db.session.commit()
        
        sync_user_to_firebase(user)
        
        return jsonify({'success': True, 'message': 'Tài khoản được thêm thành công', 'user_id': user.id})
    except Exception as e:
        db.session.rollback()
        logger.error(f"Add user error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/users/<int:user_id>', methods=['PUT'])
def update_user_api(user_id):
    """Update user via API (for admin panel user management)"""
    try:
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'message': 'Tài khoản không tồn tại'}), 404
        
        data = request.get_json()
        
        if 'student_id' in data:
            user.student_id = data['student_id']
        if 'name' in data:
            user.name = data['name']
        if 'class_name' in data:
            user.class_name = data['class_name']
        if 'khoa_vien' in data:
            user.khoa_vien = data['khoa_vien']
        if 'password' in data and data['password']:
            if len(data['password']) >= 8:
                user.set_password(data['password'])
                user.plain_password = data['password']
            else:
                return jsonify({'success': False, 'message': 'Mật khẩu phải có ít nhất 8 ký tự'}), 400
        
        db.session.commit()
        sync_user_to_firebase(user)
        
        return jsonify({'success': True, 'message': 'Cập nhật thành công'})
    except Exception as e:
        db.session.rollback()
        logger.error(f"Update user error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
def delete_user_api(user_id):
    """Delete user via API (for admin panel user management)"""
    try:
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'message': 'Tài khoản không tồn tại'}), 404
        
        Session.query.filter_by(user_id=user_id).delete()
        db.session.delete(user)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Xóa thành công'})
    except Exception as e:
        db.session.rollback()
        logger.error(f"Delete user error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/send_notification/<int:computer_id>', methods=['POST'])
def send_notification_to_computer(computer_id):
    """Send notification to a specific computer"""
    try:
        if computer_id not in client_apis:
            computer = Computer.query.get(computer_id)
            if computer and computer.ip_address:
                client_apis[computer_id] = {'ip': computer.ip_address, 'port': computer.api_port or 5001}
            else:
                return jsonify({'success': False, 'message': 'Client chưa đăng ký hoặc offline'}), 404
        
        data = request.get_json()
        title = data.get('title', 'Thông báo từ quản lý')
        message = data.get('message', '')
        notification_type = data.get('type', 'info')
        
        client_info = client_apis[computer_id]
        notify_url = f"http://{client_info['ip']}:{client_info['port']}/api/notification"
        
        response = requests.post(
            notify_url,
            json={
                'title': title,
                'message': message,
                'type': notification_type
            },
            timeout=5
        )
        
        if response.status_code == 200:
            logger.info(f"Notification sent to computer {computer_id}")
            return jsonify({'success': True, 'message': 'Gửi thông báo thành công'})
        else:
            logger.error(f"Failed to send notification to {computer_id}: {response.status_code}")
            return jsonify({'success': False, 'message': f'Không thể gửi thông báo (lỗi {response.status_code})'}), 502
            
    except requests.exceptions.Timeout:
        return jsonify({'success': False, 'message': 'Hết thời gian chờ - client không phản hồi'}), 502
    except Exception as e:
        logger.error(f"Notification error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/broadcast_notification', methods=['POST'])
def broadcast_notification():
    """Broadcast notification to all online computers"""
    try:
        data = request.get_json()
        title = data.get('title', 'Thông báo từ quản lý')
        message = data.get('message', '')
        notification_type = data.get('type', 'info')
        
        # Get all online computers
        computers = Computer.query.filter_by(status='Đang sử dụng').all()
        
        results = {'success': 0, 'failed': 0, 'errors': []}
        
        for computer in computers:
            try:
                if computer.id not in client_apis and computer.ip_address:
                    client_apis[computer.id] = {'ip': computer.ip_address, 'port': computer.api_port or 5001}
                
                if computer.id in client_apis:
                    client_info = client_apis[computer.id]
                    notify_url = f"http://{client_info['ip']}:{client_info['port']}/api/notification"
                    
                    response = requests.post(
                        notify_url,
                        json={
                            'title': title,
                            'message': message,
                            'type': notification_type
                        },
                        timeout=3
                    )
                    
                    if response.status_code == 200:
                        results['success'] += 1
                        logger.info(f"Broadcast notification sent to {computer.name}")
                    else:
                        results['failed'] += 1
                        results['errors'].append(f"{computer.name}: HTTP {response.status_code}")
            except Exception as e:
                results['failed'] += 1
                results['errors'].append(f"{computer.name}: {str(e)}")
        
        return jsonify({
            'success': True,
            'message': f'Gửi thành công: {results["success"]}, Thất bại: {results["failed"]}',
            'details': results
        })
    except Exception as e:
        logger.error(f"Broadcast notification error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# Route để render trang import_users.html (nếu chưa có)
@app.route('/admin/import_users_page')
def import_users_page():
    """Render import users page"""
    return render_template('import_users.html')

if __name__ == '__main__':
    import logging
    log = logging.getLogger('werkzeug')
    log.setLevel(logging.ERROR)
    os.environ['WERKZEUG_RUN_MAIN'] = 'false'
    
    # Suppress Flask banner when running as executable
    import io
    if sys.stdout is None:
        sys.stdout = io.StringIO()
    if sys.stderr is None:
        sys.stderr = io.StringIO()
    
    Thread(target=open_browser, daemon=True).start()
    app.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False, threaded=True)